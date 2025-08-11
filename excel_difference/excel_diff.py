import re
from copy import copy
from typing import Any, Dict, Optional, Union

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill


def is_numeric(value: Any) -> bool:
    """Check if a value is numeric."""
    if pd.isna(value):
        return False
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


def normalize_text(text: Any) -> str:
    """Normalize text for comparison by removing extra spaces and converting to lowercase."""
    if pd.isna(text):
        return ""
    return re.sub(r"\s+", " ", str(text).strip()).lower()


def find_matching_rows(
    ws1, ws2, key_column: int = 1, min_similarity: float = 0.8
) -> Dict[int, int]:
    """
    Find matching rows between two worksheets based on content similarity.

    Args:
        ws1: First worksheet
        ws2: Second worksheet
        key_column: Column to use for matching (1-based)
        min_similarity: Minimum similarity threshold for matching

    Returns:
        Dictionary mapping row numbers from ws1 to ws2
    """
    matches = {}
    used_ws2_rows = set()

    # Get all non-empty rows from both sheets
    ws1_rows = {}
    ws2_rows = {}

    # Collect rows from ws1
    for row in range(1, ws1.max_row + 1):
        cell_value = ws1.cell(row=row, column=key_column).value
        if cell_value and str(cell_value).strip():
            ws1_rows[row] = normalize_text(cell_value)

    # Collect rows from ws2
    for row in range(1, ws2.max_row + 1):
        cell_value = ws2.cell(row=row, column=key_column).value
        if cell_value and str(cell_value).strip():
            ws2_rows[row] = normalize_text(cell_value)

    print(f"Found {len(ws1_rows)} non-empty rows in file 1")
    print(f"Found {len(ws2_rows)} non-empty rows in file 2")

    # First pass: exact matches
    for row1, text1 in ws1_rows.items():
        for row2, text2 in ws2_rows.items():
            if row2 not in used_ws2_rows and text1 == text2:
                matches[row1] = row2
                used_ws2_rows.add(row2)
                print(f"Exact match: Row {row1} -> Row {row2} ({text1[:50]}...)")
                break

    # Second pass: fuzzy matches for unmatched rows
    for row1, text1 in ws1_rows.items():
        if row1 in matches:
            continue

        best_match = None
        best_similarity = 0

        for row2, text2 in ws2_rows.items():
            if row2 in used_ws2_rows:
                continue

            # Calculate similarity (simple approach)
            similarity = calculate_similarity(text1, text2)
            if similarity > best_similarity and similarity >= min_similarity:
                best_similarity = similarity
                best_match = row2

        if best_match:
            matches[row1] = best_match
            used_ws2_rows.add(best_match)
            print(
                f"Fuzzy match: Row {row1} -> Row {best_match} "
                f"(similarity: {best_similarity:.2f})"
            )

    print(f"Total matches found: {len(matches)}")
    return matches


def find_matching_columns(
    ws1, ws2, key_row: int = 1, min_similarity: float = 0.8
) -> Dict[int, int]:
    """
    Find matching columns between two worksheets based on header content similarity.

    Args:
        ws1: First worksheet
        ws2: Second worksheet
        key_row: Row to use for matching columns (1-based)
        min_similarity: Minimum similarity threshold for matching

    Returns:
        Dictionary mapping column numbers from ws1 to ws2
    """
    matches = {}
    used_ws2_cols = set()

    # Get all non-empty columns from both sheets
    ws1_cols = {}
    ws2_cols = {}

    # Collect columns from ws1
    for col in range(1, ws1.max_column + 1):
        cell_value = ws1.cell(row=key_row, column=col).value
        if cell_value and str(cell_value).strip():
            ws1_cols[col] = normalize_text(cell_value)

    # Collect columns from ws2
    for col in range(1, ws2.max_column + 1):
        cell_value = ws2.cell(row=key_row, column=col).value
        if cell_value and str(cell_value).strip():
            ws2_cols[col] = normalize_text(cell_value)

    print(f"Found {len(ws1_cols)} non-empty columns in file 1")
    print(f"Found {len(ws2_cols)} non-empty columns in file 2")

    # First pass: exact matches
    for col1, text1 in ws1_cols.items():
        for col2, text2 in ws2_cols.items():
            if col2 not in used_ws2_cols and text1 == text2:
                matches[col1] = col2
                used_ws2_cols.add(col2)
                print(f"Exact column match: Col {col1} -> Col {col2} ({text1[:50]}...)")
                break

    # Second pass: fuzzy matches for unmatched columns
    for col1, text1 in ws1_cols.items():
        if col1 in matches:
            continue

        best_match = None
        best_similarity = 0

        for col2, text2 in ws2_cols.items():
            if col2 in used_ws2_cols:
                continue

            # Calculate similarity (simple approach)
            similarity = calculate_similarity(text1, text2)
            if similarity > best_similarity and similarity >= min_similarity:
                best_similarity = similarity
                best_match = col2

        if best_match:
            matches[col1] = best_match
            used_ws2_cols.add(best_match)
            print(
                f"Fuzzy column match: Col {col1} -> Col {best_match} "
                f"(similarity: {best_similarity:.2f})"
            )

    print(f"Total column matches found: {len(matches)}")
    return matches


def calculate_similarity(text1: str, text2: str) -> float:
    """Calculate similarity between two text strings."""
    if not text1 or not text2:
        return 0.0

    # Simple word-based similarity
    words1 = set(text1.split())
    words2 = set(text2.split())

    if not words1 and not words2:
        return 1.0
    if not words1 or not words2:
        return 0.0

    intersection = words1.intersection(words2)
    union = words1.union(words2)

    return len(intersection) / len(union)


def compare_cells(value1: Any, value2: Any) -> Union[float, str]:
    """
    Compare two cell values and return the appropriate result.

    For numeric values: returns value2 - value1
    For text values: returns value1 if they match, otherwise "VALUE1 <--> VALUE2"
    """
    # Handle NaN values
    if pd.isna(value1) and pd.isna(value2):
        return 0.0
    elif pd.isna(value1):
        return value2 if is_numeric(value2) else str(value2)
    elif pd.isna(value2):
        return -value1 if is_numeric(value1) else f"{value1} <--> "

    # Check if both are numeric
    if is_numeric(value1) and is_numeric(value2):
        return float(value2) - float(value1)

    # Handle text values
    str1 = str(value1) if value1 is not None else ""
    str2 = str(value2) if value2 is not None else ""

    if str1 == str2:
        return str1
    else:
        return f"{str1} <--> {str2}"


def get_cell_color(difference_value: Any) -> Optional[PatternFill]:
    """Get color fill based on difference value."""
    if pd.isna(difference_value):
        return None

    if isinstance(difference_value, (int, float)):
        if difference_value == 0:
            return PatternFill(
                start_color="90EE90", end_color="90EE90", fill_type="solid"
            )  # Light green
        elif abs(difference_value) > 100:  # High difference
            return PatternFill(
                start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
            )  # Light red
        else:  # Medium difference
            return PatternFill(
                start_color="FFE4B5", end_color="FFE4B5", fill_type="solid"
            )  # Light orange
    elif isinstance(difference_value, str) and "<-->" in difference_value:
        return PatternFill(
            start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
        )  # Light red for text differences

    return None


def detect_key_row_column(ws) -> tuple[int, int]:
    """
    Automatically detect the best key row and column based on text content.

    Returns the row and column with the most non-numeric text values.

    Args:
        ws: Worksheet to analyze

    Returns:
        Tuple of (key_row, key_column) - 1-based indices
    """
    max_row = ws.max_row
    max_col = ws.max_column

    # Count text values in each row
    row_text_counts = {}
    for row in range(1, max_row + 1):
        text_count = 0
        for col in range(1, max_col + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and not is_numeric(cell_value):
                text_count += 1
        row_text_counts[row] = text_count

    # Count text values in each column
    col_text_counts = {}
    for col in range(1, max_col + 1):
        text_count = 0
        for row in range(1, max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and not is_numeric(cell_value):
                text_count += 1
        col_text_counts[col] = text_count

    # Find row and column with most text values
    best_row = max(row_text_counts.items(), key=lambda x: x[1])[0]
    best_col = max(col_text_counts.items(), key=lambda x: x[1])[0]

    print(
        f"Auto-detected key row: {best_row} (with {row_text_counts[best_row]} text values)"
    )
    print(
        f"Auto-detected key column: {best_col} (with {col_text_counts[best_col]} text values)"
    )

    return best_row, best_col


def excel_diff(
    file1_path: str,
    file2_path: str,
    output_path: str,
    key_column: int = 1,
    key_row: int = 1,
) -> None:
    """
    Generate a difference Excel file between two input Excel files with smart row and column matching.

    Args:
        file1_path: Path to the first Excel file
        file2_path: Path to the second Excel file
        output_path: Path for the output Excel file
        key_column: Column to use for matching rows (1-based, default=1, 0 for auto-detect)
        key_row: Row to use for matching columns (1-based, default=1, 0 for auto-detect)
    """
    print("Loading workbooks...")
    # Load both workbooks
    wb1 = openpyxl.load_workbook(file1_path)
    wb2 = openpyxl.load_workbook(file2_path)

    # Auto-detect key row and column if needed
    actual_key_row = key_row
    actual_key_column = key_column

    if key_row == 0 or key_column == 0:
        print("Auto-detecting key row and column...")
        # Use the first sheet for detection (assuming structure is similar across sheets)
        first_sheet_name = wb1.sheetnames[0] if wb1.sheetnames else None
        if first_sheet_name and first_sheet_name in wb2.sheetnames:
            ws1 = wb1[first_sheet_name]
            detected_key_row, detected_key_column = detect_key_row_column(ws1)

            if key_row == 0:
                actual_key_row = detected_key_row
            if key_column == 0:
                actual_key_column = detected_key_column
        else:
            print("Warning: Could not auto-detect key row/column, using defaults")
            if key_row == 0:
                actual_key_row = 1
            if key_column == 0:
                actual_key_column = 1

    # Create output workbook
    wb_output = openpyxl.Workbook()

    # Remove default sheet
    wb_output.remove(wb_output.active)

    # Process each sheet
    for sheet_name in wb1.sheetnames:
        if sheet_name in wb2.sheetnames:
            print(f"\nProcessing sheet: {sheet_name}")

            # Get the sheets
            ws1 = wb1[sheet_name]
            ws2 = wb2[sheet_name]

            # Create new sheet in output workbook
            ws_output = wb_output.create_sheet(title=sheet_name)

            # Find matching rows and columns
            row_matches = find_matching_rows(ws1, ws2, actual_key_column)
            column_matches = find_matching_columns(ws1, ws2, actual_key_row)

            if not row_matches:
                print(f"Warning: No matching rows found in sheet '{sheet_name}'")
                continue

            if not column_matches:
                print(f"Warning: No matching columns found in sheet '{sheet_name}'")
                continue

            # Get dimensions based on matched rows and columns
            max_col = max(ws1.max_column, ws2.max_column)

            # Add header row to show matching information
            header_row = 1
            ws_output.cell(
                row=header_row, column=1, value="Row Mapping (File1 -> File2)"
            )
            ws_output.cell(row=header_row, column=2, value="File1 Row")
            ws_output.cell(row=header_row, column=3, value="File2 Row")
            ws_output.cell(row=header_row, column=4, value="Key Value")

            # Style header
            header_font = Font(bold=True)
            for col in range(1, 5):
                ws_output.cell(row=header_row, column=col).font = header_font

            # Add row mapping information
            mapping_row = 2
            for row1, row2 in sorted(row_matches.items()):
                key_value = ws1.cell(row=row1, column=actual_key_column).value
                ws_output.cell(row=mapping_row, column=2, value=row1)
                ws_output.cell(row=mapping_row, column=3, value=row2)
                ws_output.cell(row=mapping_row, column=4, value=key_value)
                mapping_row += 1

            # Add column mapping information
            col_mapping_start = mapping_row + 2
            ws_output.cell(
                row=col_mapping_start, column=1, value="Column Mapping (File1 -> File2)"
            )
            ws_output.cell(row=col_mapping_start, column=2, value="File1 Col")
            ws_output.cell(row=col_mapping_start, column=3, value="File2 Col")
            ws_output.cell(row=col_mapping_start, column=4, value="Header Value")

            # Style column mapping header
            for col in range(1, 5):
                ws_output.cell(row=col_mapping_start, column=col).font = header_font

            col_mapping_row = col_mapping_start + 1
            for col1, col2 in sorted(column_matches.items()):
                header_value = ws1.cell(row=actual_key_row, column=col1).value
                ws_output.cell(row=col_mapping_row, column=2, value=col1)
                ws_output.cell(row=col_mapping_row, column=3, value=col2)
                ws_output.cell(row=col_mapping_row, column=4, value=header_value)
                col_mapping_row += 1

            # Add separator
            separator_row = col_mapping_row + 1
            ws_output.cell(row=separator_row, column=1, value="=" * 50)

            # Start data comparison from this row
            data_start_row = separator_row + 2

            # Process matched rows and columns
            output_row = data_start_row
            for row1, row2 in sorted(row_matches.items()):
                output_col = 1
                for col1, col2 in sorted(column_matches.items()):
                    cell1 = ws1.cell(row=row1, column=col1)
                    cell2 = ws2.cell(row=row2, column=col2)

                    # Compare values
                    result = compare_cells(cell1.value, cell2.value)

                    # Set value in output sheet
                    output_cell = ws_output.cell(row=output_row, column=output_col)
                    output_cell.value = result

                    # Apply color based on difference
                    color_fill = get_cell_color(result)
                    if color_fill:
                        output_cell.fill = color_fill

                    # Copy styling from the first workbook
                    if cell1.has_style:
                        if cell1.font:
                            output_cell.font = copy(cell1.font)
                        if cell1.border:
                            output_cell.border = copy(cell1.border)
                        if (
                            cell1.fill and not color_fill
                        ):  # Don't override difference colors
                            output_cell.fill = copy(cell1.fill)
                        if cell1.number_format:
                            output_cell.number_format = cell1.number_format
                        if cell1.protection:
                            output_cell.protection = copy(cell1.protection)
                        if cell1.alignment:
                            output_cell.alignment = copy(cell1.alignment)

                    output_col += 1

                output_row += 1

            # Copy column dimensions for matched columns
            output_col = 1
            for col1, col2 in sorted(column_matches.items()):
                if col1 in ws1.column_dimensions:
                    ws_output.column_dimensions[output_col] = ws1.column_dimensions[
                        col1
                    ]
                elif col2 in ws2.column_dimensions:
                    ws_output.column_dimensions[output_col] = ws2.column_dimensions[
                        col2
                    ]
                output_col += 1

            # Copy row dimensions for data rows
            for i, (row1, row2) in enumerate(sorted(row_matches.items())):
                output_row = data_start_row + i
                if row1 in ws1.row_dimensions:
                    ws_output.row_dimensions[output_row] = ws1.row_dimensions[row1]

            print(f"Processed {len(row_matches)} matched rows in sheet '{sheet_name}'")

    # Save the output workbook
    wb_output.save(output_path)
    print(f"\nDifference file saved to: {output_path}")
    print(
        "Note: Only matched rows are included in the output. "
        "Rows without matches are skipped."
    )


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 4:
        print(
            "Usage: python excel_diff.py <file1.xlsx> <file2.xlsx> "
            "<output.xlsx> [key_column] [key_row]"
        )
        sys.exit(1)

    file1_path = sys.argv[1]
    file2_path = sys.argv[2]
    output_path = sys.argv[3]
    key_column = int(sys.argv[4]) if len(sys.argv) > 4 else 1
    key_row = int(sys.argv[5]) if len(sys.argv) > 5 else 1

    excel_diff(file1_path, file2_path, output_path, key_column, key_row)
